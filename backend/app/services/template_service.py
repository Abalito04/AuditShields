from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.validation_service import get_entity_schema


def generate_template(entity_type: str) -> tuple[BytesIO, str]:
    schema = get_entity_schema(entity_type)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = schema.sheet_name

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for column_index, column in enumerate(schema.columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=column)
        cell.fill = header_fill
        cell.font = header_font
        sheet.cell(row=2, column=column_index, value=schema.example_row.get(column))
        sheet.column_dimensions[cell.column_letter].width = max(18, len(column) + 4)

    notes = workbook.create_sheet("notas")
    notes["A1"] = "Entidad"
    notes["B1"] = schema.label
    notes["A2"] = "Columnas obligatorias"
    notes["B2"] = ", ".join(schema.required_columns)
    notes["A3"] = "Formato de fechas"
    notes["B3"] = "YYYY-MM-DD o YYYY-MM-DD HH:MM segun corresponda"
    notes["A4"] = "Importante"
    notes["B4"] = "No cambies los nombres de columnas de la primera fila."

    for cell in notes[1]:
        cell.font = Font(bold=True)
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 70

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, f"{schema.sheet_name}.xlsx"
