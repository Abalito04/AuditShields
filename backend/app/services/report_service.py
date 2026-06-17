from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func

from app.models import Alert, Case, Supplier
from app.services.dashboard_service import get_dashboard_metrics
from app.utils.labels import CASE_STATUS_LABELS, MODULE_LABELS, RISK_LEVEL_LABELS


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def build_cases_report() -> tuple[BytesIO, str]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "casos"
    headers = [
        "case_number",
        "title",
        "module",
        "risk_level",
        "risk_score",
        "status",
        "assigned_to",
        "created_at",
        "closed_at",
        "resolution_summary",
    ]
    _write_headers(sheet, headers)
    for row_index, audit_case in enumerate(Case.query.order_by(Case.created_at.desc()).all(), start=2):
        alert = audit_case.alert
        values = [
            audit_case.case_number,
            audit_case.title,
            MODULE_LABELS.get(alert.module, alert.module),
            RISK_LEVEL_LABELS.get(audit_case.risk_level, audit_case.risk_level),
            audit_case.risk_score,
            CASE_STATUS_LABELS.get(audit_case.status, audit_case.status),
            audit_case.assigned_to.name if audit_case.assigned_to else "",
            _text(audit_case.created_at),
            _text(audit_case.closed_at),
            audit_case.resolution_summary or "",
        ]
        _write_row(sheet, row_index, values)
    _autosize(sheet)
    return _save(workbook), "casos_auditoria.xlsx"


def build_alerts_report() -> tuple[BytesIO, str]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "alertas"
    headers = [
        "rule_code",
        "rule_name",
        "module",
        "entity_type",
        "entity_id",
        "title",
        "risk_level",
        "risk_score",
        "amount_at_risk",
        "created_at",
    ]
    _write_headers(sheet, headers)
    for row_index, alert in enumerate(Alert.query.order_by(Alert.created_at.desc()).all(), start=2):
        values = [
            alert.rule.code,
            alert.rule.name,
            MODULE_LABELS.get(alert.module, alert.module),
            alert.entity_type,
            alert.entity_id,
            alert.title,
            RISK_LEVEL_LABELS.get(alert.risk_level, alert.risk_level),
            alert.risk_score,
            float(alert.amount_at_risk or 0),
            _text(alert.created_at),
        ]
        _write_row(sheet, row_index, values)
    _autosize(sheet)
    return _save(workbook), "alertas.xlsx"


def build_risk_summary_report() -> tuple[BytesIO, str]:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "resumen"
    metrics = get_dashboard_metrics()
    _write_headers(summary, ["Indicador", "Valor"])
    rows = [
        ("Alertas activas", metrics["active_alerts"]),
        ("Casos abiertos", metrics["open_cases"]),
        ("Casos criticos", metrics["critical_cases"]),
        ("Monto en riesgo", float(metrics["amount_at_risk"] or 0)),
        ("Proveedores observados", metrics["suppliers_observed"]),
        ("Productos observados", metrics["products_observed"]),
    ]
    for index, row in enumerate(rows, start=2):
        _write_row(summary, index, row)
    _autosize(summary)

    cases_by_status = workbook.create_sheet("casos_por_estado")
    _write_headers(cases_by_status, ["Estado", "Cantidad"])
    for index, (status, count) in enumerate(metrics["cases_by_status"], start=2):
        _write_row(cases_by_status, index, [CASE_STATUS_LABELS.get(status, status), count])
    _autosize(cases_by_status)

    alerts_by_rule = workbook.create_sheet("alertas_por_regla")
    _write_headers(alerts_by_rule, ["Regla", "Nombre", "Cantidad"])
    rows_by_rule = (
        Alert.query.join(Alert.rule)
        .with_entities(Alert.rule_id, func.count(Alert.id))
        .group_by(Alert.rule_id)
        .order_by(func.count(Alert.id).desc())
        .all()
    )
    for index, (rule_id, count) in enumerate(rows_by_rule, start=2):
        alert = Alert.query.filter_by(rule_id=rule_id).first()
        _write_row(alerts_by_rule, index, [alert.rule.code, alert.rule.name, count])
    _autosize(alerts_by_rule)

    amount_by_supplier = workbook.create_sheet("monto_por_proveedor")
    _write_headers(amount_by_supplier, ["Proveedor", "Monto en riesgo"])
    supplier_alerts = (
        Alert.query.filter(Alert.entity_type == "supplier")
        .with_entities(Alert.entity_id, func.coalesce(func.sum(Alert.amount_at_risk), 0))
        .group_by(Alert.entity_id)
        .order_by(func.coalesce(func.sum(Alert.amount_at_risk), 0).desc())
        .all()
    )
    for index, (supplier_id, amount) in enumerate(supplier_alerts, start=2):
        supplier = Supplier.query.get(supplier_id)
        _write_row(amount_by_supplier, index, [supplier.name if supplier else supplier_id, float(amount or 0)])
    _autosize(amount_by_supplier)

    products = workbook.create_sheet("productos_observados")
    _write_headers(products, ["Entidad", "ID", "Titulo", "Riesgo", "Monto"])
    product_alerts = Alert.query.filter(
        Alert.entity_type.in_(["product", "inventory_snapshot", "stock_movement"])
    ).order_by(Alert.created_at.desc()).all()
    for index, alert in enumerate(product_alerts, start=2):
        _write_row(
            products,
            index,
            [
                alert.entity_type,
                alert.entity_id,
                alert.title,
                RISK_LEVEL_LABELS.get(alert.risk_level, alert.risk_level),
                float(alert.amount_at_risk or 0),
            ],
        )
    _autosize(products)

    return _save(workbook), "resumen_riesgo.xlsx"


def _write_headers(sheet, headers):
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _write_row(sheet, row_index: int, values):
    for column_index, value in enumerate(values, start=1):
        sheet.cell(row=row_index, column=column_index, value=value)


def _autosize(sheet):
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)


def _save(workbook: Workbook) -> BytesIO:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _text(value) -> str:
    return value.isoformat() if value else ""
